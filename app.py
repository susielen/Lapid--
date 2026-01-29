import streamlit as st
import pandas as pd
import io, re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Conciliador Contábil Pro", layout="wide")
st.title("🤖 Conciliador: Identificação da Empresa e Layout Final")

arquivo = st.file_uploader("Suba o Razão aqui", type=["csv", "xlsx"])

def extrair_nf(t):
    m = re.search(r'(?:NFE|NF|NOTA|Nº)\s*(\d+)', str(t).upper())
    if m:
        try: return int(m.group(1))
        except: return m.group(1)
    return ""

def limpar_nome_correto(l):
    l = str(l).replace('nan', '').replace('NAN', '').upper()
    match_cod = re.search(r'CONTA:\s*(\d+)', l)
    codigo = match_cod.group(1) if match_cod else ""
    nome = l.split("CONTA:")[-1].replace('NOME:', '').strip()
    nome = re.sub(r'(\d+\.)+\d+', '', nome) 
    nome = nome.replace(codigo, '').strip()
    nome = re.sub(r'^[ \-_]+', '', nome)
    return f"{codigo} - {nome}" if codigo else nome

if arquivo is not None:
    try:
        df_raw = pd.read_excel(arquivo) if arquivo.name.endswith('.xlsx') else pd.read_csv(arquivo, encoding='latin-1', sep=None, engine='python')
        
        # BUSCANDO DADOS DA EMPRESA (Geralmente nas primeiras linhas do Razão)
        nome_empresa = "EMPRESA NÃO IDENTIFICADA"
        cnpj_empresa = "CNPJ NÃO IDENTIFICADO"
        
        # Procura nas primeiras 10 linhas do arquivo original
        for i in range(min(10, len(df_raw))):
            linha_cabecalho = " ".join([str(v) for v in df_raw.iloc[i].values]).upper()
            if "EMPRESA:" in linha_cabecalho or i == 0: # Tenta pegar a primeira linha
                if not "CONTA:" in linha_cabecalho:
                    nome_empresa = linha_cabecalho.split("EMPRESA:")[-1].split("CNPJ:")[0].strip()
            if "CNPJ:" in linha_cabecalho:
                match_cnpj = re.search(r'(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})|(\d{14})', linha_cabecalho)
                if match_cnpj:
                    cnpj_empresa = f"CNPJ: {match_cnpj.group(0)}"

        resumo = {}
        fornecedor = None
        
        for i, linha in df_raw.iterrows():
            texto = " ".join([str(v) for v in linha.values]).upper()
            if "CONTA:" in texto:
                fornecedor = limpar_nome_correto(texto)
                resumo[fornecedor] = []
            elif fornecedor and ("/" in str(linha.iloc[0]) or "-" in str(linha.iloc[0])):
                def n(v):
                    v = str(v).replace('.', '').replace(',', '.')
                    try: return float(v)
                    except: return 0.0
                d, c = n(linha.iloc[8]), n(linha.iloc[9])
                if d > 0 or c > 0:
                    h = str(linha.iloc[2]).replace('nan', '')
                    try: dt_obj = pd.to_datetime(linha.iloc[0], dayfirst=True)
                    except: dt_obj = str(linha.iloc[0])
                    resumo[fornecedor].append({'Data': dt_obj, 'Nº NF': extrair_nf(h), 'Histórico': h, 'Débito': d, 'Crédito': c})

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for nome_forn, lancs in resumo.items():
                if not lancs: continue
                df_f = pd.DataFrame(lancs)
                df_c = df_f.groupby('Nº NF').agg({'Débito': 'sum', 'Crédito': 'sum'}).reset_index()
                df_c['DIFERENÇA'] = df_c['Crédito'] - df_c['Débito']
                df_c['STATUS'] = df_c['DIFERENÇA'].apply(lambda x: "OK" if abs(x) < 0.05 else "DIVERGENTE")
                
                aba = re.sub(r'[\\/*?:\[\]]', '', nome_forn)[:31]
                # Empurrei o início dos dados para a linha 10 para caber o cabeçalho novo
                df_f.to_excel(writer, sheet_name=aba, index=False, startrow=9, startcol=1)
                df_c.to_excel(writer, sheet_name=aba, index=False, startrow=9, startcol=9)
                ws = writer.sheets[aba]
                
                fundo_br = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                for r in range(1, 201):
                    for c in range(1, 21):
                        ws.cell(row=r, column=c).fill = fundo_br

                b = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                ac, ar = Alignment(horizontal='center'), Alignment(horizontal='right')
                fmt_moeda = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'

                # 1. CABEÇALHO DA SUA EMPRESA (Linhas 2 e 3)
                ws.merge_cells('B2:N2')
                ws['B2'] = nome_empresa
                ws['B2'].font = Font(bold=True, size=12, color="4F81BD")
                ws['B2'].alignment = ac

                ws.merge_cells('B3:N3')
                ws['B3'] = cnpj_empresa
                ws['B3'].font = Font(bold=False, size=10)
                ws['B3'].alignment = ac

                # 2. NOME DO FORNECEDOR (Linha 5)
                ws.merge_cells('B5:N5')
                ws['B5'] = f"CONCILIAÇÃO: {nome_forn}"
                ws['B5'].font, ws['B5'].alignment = Font(bold=True, size=14), ac

                # 3. SALDO (Linha 6)
                ws.cell(row=6, column=5, value="SALDO").alignment = ar
                val_s = df_f['Crédito'].sum() - df_f['Débito'].sum()
                cs = ws.cell(row=6, column=6, value=val_s)
                cs.font = Font(bold=True, color="FF0000" if val_s < 0 else "00B050")
                cs.border, cs.number_format = b, fmt_moeda

                # 4. TOTAIS E CONCILIAÇÃO (Linha 8)
                ws.cell(row=8, column=4, value="TOTAIS").alignment = ar
                for ci, val, cor in [(5, df_f['Débito'].sum(), "FF0000"), (6, df_f['Crédito'].sum(), "00B050")]:
                    cel = ws.cell(row=8, column=ci, value=val)
                    cel.font, cel.border, cel.number_format = Font(bold=True, color=cor), b, fmt_moeda

                ws.merge_cells('J8:L8')
                ws['J8'] = "CONCILIAÇÃO POR NOTA"
                ws['J8'].font, ws['J8'].alignment = Font(bold=True), ac
                ws.cell(row=8, column=13, value="Saldo").font = Font(bold=True)
                cc = ws.cell(row=8, column=14, value=val_s)
                cc.font, cc.border, cc.number_format = Font(bold=True, color="FF0000" if val_s < 0 else "00B050"), b, fmt_moeda

                # 5. CABEÇALHOS DAS TABELAS (LINHA 10)
                cinza = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                for c in range(2, 15):
                    cel = ws.cell(row=10, column=c)
                    if cel.value:
                        cel.font, cel.alignment, cel.fill = Font(bold=True), ac, cinza
                        if c != 7: cel.border = b

                # 6. CORPO
                for r in range(11, 11 + len(df_f)):
                    for c in range(2, 8):
                        cel = ws.cell(row=r, column=c)
                        if c < 7: cel.border = b
                        if c == 2: cel.number_format = 'dd/mm/yyyy'
                        if c in [2, 3]: cel.alignment = ac
                        if c == 6: cel.font = Font(color="FF0000")
                        if c == 7: cel.font = Font(color="00B050")
                        if c in [6, 7]: cel.number_format = fmt_moeda

                for r in range(11, 11 + len(df_c)):
                    for c in range(10, 15):
                        cel = ws.cell(row=r, column=c)
                        cel.border = b
                        if c == 10: cel.alignment = ac
                        if c in [11, 12, 13]: cel.number_format = fmt_moeda
                    st_c = ws.cell(row=r, column=14)
                    st_c.font = Font(color="00B050") if st_c.value == "OK" else Font(color="FF0000")

                # 7. AJUSTES FINAIS
                ws.column_dimensions['A'].width = 2
                ws.row_dimensions[1].height = 10
                for c in range(2, 15):
                    letra = get_column_letter(c)
                    if letra in ['G', 'H', 'I']: ws.column_dimensions[letra].width = 4
                    elif letra == 'D': ws.column_dimensions[letra].width = 45
                    else: ws.column_dimensions[letra].width = 18

        st.success("✅ Relatório completo com dados da empresa!")
        st.download_button("📥 Baixar Excel", output.getvalue(), "conciliacao_empresa.xlsx")
    except Exception as e:
        st.error(f"Erro: {e}")
